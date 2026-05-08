"""
Mock data generator for the Financial Consolidation Engine.

Run this script first to create all test data files:
    python mock_data_generator.py

Generates:
  - data/coa_mapping.xlsx
  - data/intercompany_matrix.xlsx
  - data/subsidiaries/parent_co.xlsx
  - data/subsidiaries/sub_sgd.xlsx
  - data/subsidiaries/sub_usd.xlsx
  - data/subsidiaries/sub_eur.xlsx

Design notes:
  - FX rates (from config/fx_rates.json): SGD closing=0.7407, avg=0.7350, hist=0.7200
                                           EUR closing=1.0850, avg=1.0820, hist=1.0500
  - Intercompany transactions:
      Sub SGD ↔ Parent Co  : IC Receivable/Payable 50,000 SGD   (BS elimination)
      Sub USD ↔ Sub EUR    : IC Revenue/COGS         30,000 USD  (IS elimination)
      Parent Co ↔ Sub SGD  : Intercompany Loan       100,000 USD (BS elimination)
  - Sub SGD has local account "999" (Miscellaneous) NOT in COA → triggers unmapped warning
  - Sub USD is 80% owned → NCI triggered
  - All entity BS must balance in LOCAL currency (assets = liabilities + equity)
"""

from pathlib import Path
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill

BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"
SUBS_DIR = DATA_DIR / "subsidiaries"


def main():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    SUBS_DIR.mkdir(parents=True, exist_ok=True)

    write_coa_mapping()
    write_intercompany_matrix()
    write_parent_co()
    write_sub_sgd()
    write_sub_usd()
    write_sub_eur()
    print("Mock data generated successfully.")


# ---------------------------------------------------------------------------
# COA Mapping
# ---------------------------------------------------------------------------

COA_ROWS = [
    # group_code, group_account_name, sub_sgd_code, sub_usd_code, sub_eur_code, parent_co_code, classification
    ("1000", "Cash and Cash Equivalents",      "101",    "CASH",   "C1000",  "1000",  "Asset"),
    ("1100", "Trade Receivables",              "110",    "AR",     "1100",   "1100",  "Asset"),
    ("1200", "Intercompany Receivables",       "120",    "IC-REC", "1200",   "1200",  "Asset"),
    ("1300", "Inventory",                      "130",    "INV",    "1300",   "1300",  "Asset"),
    ("1400", "Prepayments",                    "140",    "PREP",   "1400",   "1400",  "Asset"),
    ("1500", "Property, Plant & Equipment",    "150",    "PPE",    "1500",   "1500",  "Asset"),
    ("1600", "Intangible Assets",              "160",    "INTA",   "1600",   "1600",  "Asset"),
    ("1700", "Investment in Subsidiaries",     "",       "",       "",       "1700",  "Asset"),
    ("2000", "Trade Payables",                 "200",    "AP",     "2000",   "2000",  "Liability"),
    ("2100", "Intercompany Payables",          "210",    "IC-PAY", "2100",   "2100",  "Liability"),
    ("2200", "Short-term Borrowings",          "220",    "STB",    "2200",   "2200",  "Liability"),
    ("2300", "Long-term Borrowings",           "230",    "LTB",    "2300",   "2300",  "Liability"),
    ("2400", "Deferred Tax Liability",         "240",    "DTL",    "2400",   "2400",  "Liability"),
    ("2500", "Other Liabilities",              "250",    "OTL",    "2500",   "2500",  "Liability"),
    ("3000", "Share Capital",                  "SC",     "CAP",    "3000",   "3000",  "Equity"),
    ("3100", "Retained Earnings",              "RE",     "RET",    "3100",   "3100",  "Equity"),
    ("3200", "Other Comprehensive Income",     "OCI",    "OCI",    "3200",   "3200",  "Equity"),
    ("3300", "Non-Controlling Interest",       "",       "",       "",       "",      "Equity"),
    ("4000", "Revenue",                        "400",    "REV",    "4000",   "4000",  "Income"),
    ("4100", "Intercompany Revenue",           "410",    "IC-REV", "4100",   "4100",  "Income"),
    ("5000", "Cost of Goods Sold",             "500",    "COGS",   "5000",   "5000",  "Expense"),
    ("5100", "Intercompany COGS",              "510",    "IC-CGS", "5100",   "5100",  "Expense"),
    ("6000", "Gross Profit",                   "",       "",       "",       "",      "Income"),
    ("6100", "Operating Expenses",             "610",    "OPEX",   "6100",   "6100",  "Expense"),
    ("6200", "Depreciation & Amortisation",    "620",    "DEP",    "6200",   "6200",  "Expense"),
    ("6300", "Finance Income",                 "630",    "FINC",   "6300",   "6300",  "Income"),
    ("6400", "Finance Costs",                  "640",    "FCST",   "6400",   "6400",  "Expense"),
    ("6500", "Income Tax Expense",             "650",    "TAX",    "6500",   "6500",  "Expense"),
    ("6600", "Net Profit",                     "",       "",       "",       "",      "Income"),
    ("7000", "Net Profit (from IS)",           "700",    "NP",     "7000",   "7000",  "CF"),
    ("7100", "Add: Depreciation & Amort.",     "710",    "DA",     "7100",   "7100",  "CF"),
    # Note: sub_usd_code "DA" above is for CF (add-back of D&A). IS depreciation uses "DEP".
    ("7200", "Changes in Working Capital",     "720",    "WC",     "7200",   "7200",  "CF"),
    ("7300", "Cash from Operating Activities", "",       "",       "",       "",      "CF"),
    ("7400", "Purchase of PPE",                "740",    "CAPEX",  "7400",   "7400",  "CF"),
    ("7500", "Cash from Investing Activities", "",       "",       "",       "",      "CF"),
    ("7600", "Proceeds from Borrowings",       "760",    "BORP",   "7600",   "7600",  "CF"),
    ("7700", "Repayment of Borrowings",        "770",    "BORR",   "7700",   "7700",  "CF"),
    ("7800", "Dividends Paid",                 "780",    "DIV",    "7800",   "7800",  "CF"),
    ("7900", "Cash from Financing Activities", "",       "",       "",       "",      "CF"),
    ("8000", "Net Change in Cash",             "",       "",       "",       "",      "CF"),
]


def write_coa_mapping():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "COA Mapping"
    headers = ["group_code", "group_account_name", "sub_sgd_code", "sub_usd_code", "sub_eur_code", "parent_co_code", "classification"]
    ws.append(headers)
    for row in COA_ROWS:
        ws.append(list(row))
    _bold_header(ws)
    wb.save(DATA_DIR / "coa_mapping.xlsx")
    print("  coa_mapping.xlsx created")


# ---------------------------------------------------------------------------
# Intercompany Matrix
# ---------------------------------------------------------------------------

# Override-only rows: relationships NOT covered by any ICO tab.
# The Trade Receivable and IC Loan pairs are now declared in ICO tabs.
# Only the Dividend arrangement (approved manual override) remains here.
# account_type must be identical on both sides for bilateral matching.
ICO_MATRIX_OVERRIDE_ROWS = [
    # from_entity, to_entity, account_type, group_account_code, amount_original_ccy, currency, side
    ("Parent Co", "Sub SGD", "Dividend", "2500", 10000, "USD", "liability"),
    ("Sub SGD",   "Parent Co", "Dividend", "1100", 10000, "USD", "asset"),
]


def write_intercompany_matrix():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ICO Matrix"
    headers = ["from_entity", "to_entity", "account_type", "group_account_code", "amount_original_ccy", "currency", "side"]
    ws.append(headers)
    for row in ICO_MATRIX_OVERRIDE_ROWS:
        ws.append(list(row))
    _bold_header(ws)
    wb.save(DATA_DIR / "intercompany_matrix.xlsx")
    print("  intercompany_matrix.xlsx created")


# ---------------------------------------------------------------------------
# Parent Co (USD)
# Intercompany items:
#   1200: 0 (parent has no IC receivable in this exercise)
#   1700: Investment in Subsidiaries = 375,000 USD
#         (Sub SGD share capital: 500,000 SGD × hist 0.72 = 360,000 USD)
#         (Sub USD share capital: 100,000 USD × 80% parent share... but elimination is full SC)
#         (Sub EUR share capital: 200,000 EUR × hist 1.05 = 210,000 USD)
#         Simplified: 375,000 total — see note in eliminations engine about residual
#   2100: IC Payable to Sub SGD = 50,000 SGD × closing 0.7407 = 37,035 USD
#   2300: Includes 100,000 USD interco loan to Sub SGD
# BS must balance (Assets = Liabilities + Equity)
# ---------------------------------------------------------------------------

def write_parent_co():
    # Balance design (USD):
    # Interco loan: Parent lent 100,000 SGD to Sub SGD.
    #   Parent 1200 = 100,000 SGD x closing 0.7407 = 74,070 USD (IC loan receivable)
    # Investment in Subsidiaries = ownership% x sub SC at historical rates:
    #   Sub SGD: 100% x 500,000 SGD x hist 0.72 = 360,000
    #   Sub USD:  80% x 100,000 USD x hist 1.00 =  80,000  (NCI holds other 20%)
    #   Sub EUR: 100% x 200,000 EUR x hist 1.05 = 210,000
    #   Total 1700 = 650,000
    #
    # ASSETS:
    #   1000 Cash: 200,000
    #   1100 Trade Receivables: 150,000
    #   1200 IC Loan Receivable (Sub SGD): 74,070
    #   1700 Investment in Subsidiaries: 650,000
    #   1500 PPE: 200,000
    #   Total Assets = 1,274,070
    #
    # LIABILITIES:
    #   2100 IC Payable (to Sub SGD 50k SGD trade): 37,035
    #   2300 LT Borrowings (external only): 150,000
    #   2000 Trade Payables: 100,000
    #   Total Liabilities = 287,035
    #
    # EQUITY:
    #   3000 Share Capital: 500,000
    #   3100 Retained Earnings: 487,035
    #   Total Equity = 987,035
    #
    # Check: 287,035 + 987,035 = 1,274,070 ✓

    bs = [
        ("1000", "Cash and Cash Equivalents",    200000),
        ("1100", "Trade Receivables",            150000),
        ("1200", "IC Loan Receivable (Sub SGD)",  74070),
        ("1700", "Investment in Subsidiaries",   650000),
        ("1500", "Property, Plant & Equipment",  200000),
        ("2000", "Trade Payables",               100000),
        ("2100", "Intercompany Payables",         37035),
        ("2300", "Long-term Borrowings",         150000),
        ("3000", "Share Capital",                500000),
        ("3100", "Retained Earnings",            487035),
    ]

    # IS (USD):
    # Revenue: 800,000; COGS: 400,000; OpEx: 100,000; D&A: 30,000
    # Finance Income: 10,000; Finance Costs: 15,000; Tax: 57,000; Net Profit: 208,000
    is_ = [
        ("4000", "Revenue",                    800000),
        ("5000", "Cost of Goods Sold",         400000),
        ("6100", "Operating Expenses",         100000),
        ("6200", "Depreciation & Amortisation", 30000),
        ("6300", "Finance Income",              10000),
        ("6400", "Finance Costs",               15000),
        ("6500", "Income Tax Expense",          57000),
    ]

    # CF (USD):
    cf = [
        ("7000", "Net Profit (from IS)",            208000),
        ("7100", "Add: Depreciation & Amort.",       30000),
        ("7200", "Changes in Working Capital",      -20000),
        ("7400", "Purchase of PPE",                 -50000),
        ("7600", "Proceeds from Borrowings",         20000),
        ("7700", "Repayment of Borrowings",         -10000),
        ("7800", "Dividends Paid",                  -30000),
    ]

    # EQ (USD):
    # Net Profit: 208,000; Dividends: -30,000; net movement = 178,000
    # Closing RE: 487,035; Opening RE = 487,035 - 178,000 = 309,035
    eq = [
        ("Share Capital",              500000, 0,       500000),
        ("Retained Earnings",          309035, 178000,  487035),
        ("Other Comprehensive Income",      0,      0,       0),
    ]

    # ICO tab: local account codes matching parent_co_code column in COA
    # Parent Co BS shows 2100 IC Payable to Sub SGD for trade (Parent owes Sub SGD)
    # and 1200 IC Loan Receivable from Sub SGD (Parent lent to Sub SGD)
    parent_ico = [
        # account_no, to_party, account_type, amount, currency, description
        ("2100", "Sub SGD", "Trade Receivable", 50000,  "SGD", "Q4 goods"),
        ("1200", "Sub SGD", "IC Loan",         100000, "SGD", "Facility drawdown"),
    ]

    _write_subsidiary(
        SUBS_DIR / "parent_co.xlsx",
        entity_name="Parent Co",
        currency="USD",
        reporting_period="2024-12",
        ownership_pct=100,
        bs_rows=bs,
        is_rows=is_,
        cf_rows=cf,
        eq_rows=eq,
        local_col="parent_co_code",
        ico_rows=parent_ico,
    )
    print("  parent_co.xlsx created")


# ---------------------------------------------------------------------------
# Sub SGD (SGD, 100% owned)
# Intercompany items:
#   1200: IC Receivable = 50,000 SGD (from Parent Co)
#   2300: IC Loan from Parent = 100,000 USD → recorded by Sub SGD as 100,000 / 0.7407 = ~135,014 SGD
#   Account "999" Miscellaneous: 5,000 SGD → NOT in COA → unmapped warning
# BS must balance in SGD
# ---------------------------------------------------------------------------

def write_sub_sgd():
    # IC loan from parent = 100,000 SGD (Parent lent SGD directly).
    # Recorded in Sub SGD at face value: 100,000 SGD.
    # At closing rate 0.7407 = 74,070 USD — matches Parent's 1200 exactly.
    ic_loan_sgd = 100000  # SGD (clean round number)

    # ASSETS:
    #   1000 Cash: 100,000
    #   1100 Trade Receivables: 80,000
    #   1200 IC Receivables: 50,000    ← ICO (trade receivable from Parent Co)
    #   1300 Inventory: 60,000
    #   1500 PPE: 200,000
    #   Total Assets = 490,000
    #
    # LIABILITIES:
    #   2000 Trade Payables: 40,000
    #   2300 LT Borrowings (IC loan from Parent): 100,000 SGD
    #   Total Liabilities = 140,000
    #
    # EQUITY:
    #   3000 Share Capital: 500,000  → at hist rate 0.72 = 360,000 USD
    #   3100 Retained Earnings: -150,000  → to make BS balance
    #   Total Equity = 350,000
    #
    # Check: 140,000 + 350,000 = 490,000 ✓

    total_assets = 490000
    sc = 500000
    liabilities = 40000 + ic_loan_sgd
    re = total_assets - liabilities - sc
    # re = 490,000 - 140,000 - 500,000 = -150,000

    bs = [
        ("101",  "Cash and Cash Equivalents",    100000),
        ("110",  "Trade Receivables",             80000),
        ("120",  "Intercompany Receivables",      50000),
        ("130",  "Inventory",                     60000),
        ("150",  "Property, Plant & Equipment",  200000),
        ("200",  "Trade Payables",                40000),
        ("230",  "Long-term Borrowings",    ic_loan_sgd),
        ("SC",   "Share Capital",                500000),
        ("RE",   "Retained Earnings",                re),
        ("999",  "Miscellaneous",                 5000),  # INTENTIONALLY UNMAPPED
    ]

    # IS (SGD):
    # Revenue: 500,000
    # COGS: 300,000
    # OpEx: 80,000
    # D&A: 20,000
    # Finance Costs: 5,000
    # Tax: 28,750
    # Net Profit: 66,250
    is_ = [
        ("400", "Revenue",                    500000),
        ("500", "Cost of Goods Sold",         300000),
        ("610", "Operating Expenses",          80000),
        ("620", "Depreciation & Amortisation", 20000),
        ("640", "Finance Costs",                5000),
        ("650", "Income Tax Expense",          28750),
        ("650b","Net Profit",                  66250),  # local code for Net Profit (will be mapped separately)
    ]
    # Note: local code for Net Profit doesn't exist in COA as a line item normally,
    # but we add it here for completeness. The COA has 6600="Net Profit" with no sub_sgd_code → unmapped.
    # Let's use a proper mapping: we'll add local code "NP" for Net Profit... actually per COA_ROWS
    # group 6600 "Net Profit" has sub_sgd_code="" so any local code we put won't map.
    # Better: omit Net Profit from IS input (it's a calculated line). Use code that won't map.
    # The consolidator calculates it from Revenue - expenses.
    # Let's just leave it out — calculated lines have no sub code.
    is_ = [
        ("400", "Revenue",                    500000),
        ("500", "Cost of Goods Sold",         300000),
        ("610", "Operating Expenses",          80000),
        ("620", "Depreciation & Amortisation", 20000),
        ("640", "Finance Costs",                5000),
        ("650", "Income Tax Expense",          28750),
    ]

    cf = [
        ("700", "Net Profit (from IS)",            66250),
        ("710", "Add: Depreciation & Amort.",       20000),
        ("720", "Changes in Working Capital",      -10000),
        ("740", "Purchase of PPE",                 -30000),
        ("760", "Proceeds from Borrowings",       100000),  # IC loan from Parent (SGD)
        ("780", "Dividends Paid",                  -15000),
    ]

    # EQ (SGD): Net movement = NP 66,250 - dividends 15,000 = 51,250
    # Closing RE = -150,000; Opening RE = -150,000 - 51,250 = -201,250
    opening_re = re - (66250 - 15000)
    eq = [
        ("Share Capital",              500000, 0,      500000),
        ("Retained Earnings",    opening_re, 51250,        re),
        ("Other Comprehensive Income",      0,     0,       0),
    ]

    # ICO tab: local account codes matching sub_sgd_code column in COA
    # 120 = Intercompany Receivables, 230 = Long-term Borrowings
    sgd_ico = [
        # account_no, to_party, account_type, amount, currency, description
        ("120", "Parent Co", "Trade Receivable", 50000,  "SGD", "Q4 goods"),
        ("230", "Parent Co", "IC Loan",         100000, "SGD", "Facility drawdown"),
    ]

    _write_subsidiary(
        SUBS_DIR / "sub_sgd.xlsx",
        entity_name="Sub SGD",
        currency="SGD",
        reporting_period="2024-12",
        ownership_pct=100,
        bs_rows=bs,
        is_rows=is_,
        cf_rows=cf,
        eq_rows=eq,
        local_col="sub_sgd_code",
        ico_rows=sgd_ico,
    )
    print("  sub_sgd.xlsx created")


# ---------------------------------------------------------------------------
# Sub USD (USD, 80% owned → NCI 20%)
# Intercompany:
#   4100: IC Revenue = 30,000 USD (from Sub EUR)
# Net Profit designed = 120,000 USD → NCI P&L = 24,000 USD
# Net Assets (equity) designed ≈ 600,000 USD → NCI BS = 120,000 USD
# BS must balance in USD
# ---------------------------------------------------------------------------

def write_sub_usd():
    # ASSETS:
    #   1000 Cash: 200,000
    #   1100 Trade Receivables: 150,000
    #   1300 Inventory: 100,000
    #   1500 PPE: 350,000
    #   Total Assets = 800,000
    #
    # LIABILITIES:
    #   2000 Trade Payables: 80,000
    #   2200 Short-term Borrowings: 120,000
    #   Total Liabilities = 200,000
    #
    # EQUITY:
    #   3000 Share Capital: 100,000
    #   3100 Retained Earnings: 500,000
    #   Total Equity = 600,000
    #   Net Assets = 600,000 → NCI = 20% × 600,000 = 120,000 ✓
    #
    # Check: 200,000 + 600,000 = 800,000 ✓

    bs = [
        ("CASH",   "Cash and Cash Equivalents",   200000),
        ("AR",     "Trade Receivables",            150000),
        ("INV",    "Inventory",                    100000),
        ("PPE",    "Property, Plant & Equipment",  350000),
        ("AP",     "Trade Payables",                80000),
        ("STB",    "Short-term Borrowings",        120000),
        ("CAP",    "Share Capital",                100000),
        ("RET",    "Retained Earnings",            500000),
    ]

    # IS (USD):
    # Revenue: 500,000 (external)
    # IC Revenue: 30,000
    # COGS: 300,000
    # OpEx: 60,000
    # D&A: 20,000
    # Finance Costs: 5,000
    # Tax: 25,000
    # Net Profit = (500,000 + 30,000) - 300,000 - 60,000 - 20,000 - 5,000 - 25,000 = 120,000 ✓
    is_ = [
        ("REV",    "Revenue",                     500000),
        ("IC-REV", "Intercompany Revenue",          30000),
        ("COGS",   "Cost of Goods Sold",           300000),
        ("OPEX",   "Operating Expenses",            60000),
        ("DEP",    "Depreciation & Amortisation",   20000),  # "DEP" -> COA 6200
        ("FCST",   "Finance Costs",                  5000),
        ("TAX",    "Income Tax Expense",            25000),
    ]

    cf = [
        ("NP",     "Net Profit (from IS)",         120000),
        ("DA",     "Add: Depreciation & Amort.",    20000),  # "DA" -> COA 7100
        ("WC",     "Changes in Working Capital",   -15000),
        ("CAPEX",  "Purchase of PPE",              -40000),
        ("DIV",    "Dividends Paid",               -20000),
    ]

    # EQ (USD): Opening RE = 500,000 - (120,000 - 20,000) = 400,000
    eq = [
        ("Share Capital",              100000,       0, 100000),
        ("Retained Earnings",          400000,  100000, 500000),
        ("Other Comprehensive Income",      0,       0,      0),
    ]

    # ICO tab: local account codes matching sub_usd_code column in COA
    # IC-REV = Intercompany Revenue (sub_usd_code)
    usd_ico = [
        # account_no, to_party, account_type, amount, currency, description
        ("IC-REV", "Sub EUR", "IC Revenue/COGS", 30000, "USD", "Services Q4"),
    ]

    _write_subsidiary(
        SUBS_DIR / "sub_usd.xlsx",
        entity_name="Sub USD",
        currency="USD",
        reporting_period="2024-12",
        ownership_pct=80,
        bs_rows=bs,
        is_rows=is_,
        cf_rows=cf,
        eq_rows=eq,
        local_col="sub_usd_code",
        ico_rows=usd_ico,
    )
    print("  sub_usd.xlsx created")


# ---------------------------------------------------------------------------
# Sub EUR (EUR, 100% owned)
# Intercompany:
#   5100: IC COGS = 27,726 EUR (= 30,000 USD / 1.0820 average rate)
#         After translation: 27,726 × 1.0820 ≈ 29,999.77 USD ≈ 30,000 USD (within 1 USD tolerance)
# BS must balance in EUR
# ---------------------------------------------------------------------------

def write_sub_eur():
    ic_cogs_eur = round(30000 / 1.0820, 0)  # ≈ 27,726 EUR

    # ASSETS:
    #   1000 Cash: 80,000
    #   1100 Trade Receivables: 100,000
    #   1500 PPE: 300,000
    #   Total Assets = 480,000
    #
    # LIABILITIES:
    #   2000 Trade Payables: 70,000
    #   2300 LT Borrowings: 10,000
    #   Total Liabilities = 80,000
    #
    # EQUITY:
    #   3000 Share Capital: 200,000  → at hist rate 1.05 = 210,000 USD
    #   3100 Retained Earnings: 200,000
    #   Total Equity = 400,000
    #
    # Check: 80,000 + 400,000 = 480,000 ✓

    bs = [
        ("C1000", "Cash and Cash Equivalents",    80000),
        ("1100",  "Trade Receivables",           100000),
        ("1500",  "Property, Plant & Equipment", 300000),
        ("2000",  "Trade Payables",               70000),
        ("2300",  "Long-term Borrowings",         10000),
        ("3000",  "Share Capital",               200000),
        ("3100",  "Retained Earnings",           200000),
    ]

    # IS (EUR):
    # Revenue: 300,000
    # IC COGS: 27,726 EUR
    # COGS (external): 180,000
    # OpEx: 40,000
    # D&A: 15,000
    # Finance Costs: 3,000
    # Tax: 15,069  (to get clean Net Profit)
    # Net Profit = 300,000 - 27,726 - 180,000 - 40,000 - 15,000 - 3,000 - 15,069 = 19,205 EUR
    net_profit_eur = round(300000 - ic_cogs_eur - 180000 - 40000 - 15000 - 3000 - 15069, 0)

    is_ = [
        ("4000",  "Revenue",                     300000),
        ("5100",  "Intercompany COGS",       ic_cogs_eur),
        ("5000",  "Cost of Goods Sold",          180000),
        ("6100",  "Operating Expenses",           40000),
        ("6200",  "Depreciation & Amortisation",  15000),
        ("6400",  "Finance Costs",                 3000),
        ("6500",  "Income Tax Expense",           15069),
    ]

    cf = [
        ("7000",  "Net Profit (from IS)",     net_profit_eur),
        ("7100",  "Add: Depreciation & Amort.",   15000),
        ("7200",  "Changes in Working Capital",  -10000),
        ("7400",  "Purchase of PPE",             -25000),
        ("7600",  "Proceeds from Borrowings",     10000),
        ("7800",  "Dividends Paid",              -10000),
    ]

    # EQ (EUR): Opening RE = 200,000 - (net_profit_eur - 10,000 dividends)
    opening_re = round(200000 - (net_profit_eur - 10000), 0)
    eq = [
        ("Share Capital",              200000,            0, 200000),
        ("Retained Earnings",    opening_re, net_profit_eur - 10000, 200000),
        ("Other Comprehensive Income",      0,            0,       0),
    ]

    # ICO tab: local account codes matching sub_eur_code column in COA
    # 5100 = Intercompany COGS (sub_eur_code)
    # EUR amount = 30,000 USD / 1.0820 average rate ≈ 27,726 EUR
    eur_ico = [
        # account_no, to_party, account_type, amount, currency, description
        ("5100", "Sub USD", "IC Revenue/COGS", round(30000 / 1.0820, 0), "EUR", "Services Q4"),
    ]

    _write_subsidiary(
        SUBS_DIR / "sub_eur.xlsx",
        entity_name="Sub EUR",
        currency="EUR",
        reporting_period="2024-12",
        ownership_pct=100,
        bs_rows=bs,
        is_rows=is_,
        cf_rows=cf,
        eq_rows=eq,
        local_col="sub_eur_code",
        ico_rows=eur_ico,
    )
    print("  sub_eur.xlsx created")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _write_subsidiary(
    path: Path,
    entity_name: str,
    currency: str,
    reporting_period: str,
    ownership_pct: float,
    bs_rows: list,
    is_rows: list,
    cf_rows: list,
    eq_rows: list,
    local_col: str,
    ico_rows: Optional[list] = None,
):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_statement_sheet(wb, "BS", bs_rows)
    _write_statement_sheet(wb, "IS", is_rows)
    _write_statement_sheet(wb, "CF", cf_rows)
    _write_eq_sheet(wb, "EQ", eq_rows)
    _write_meta_sheet(wb, entity_name, currency, reporting_period, ownership_pct)
    _write_ico_sheet(wb, ico_rows or [])

    wb.save(path)


def _write_statement_sheet(wb, tab_name: str, rows: list):
    ws = wb.create_sheet(tab_name)
    ws.append(["account_code", "account_name", "amount"])
    _bold_header(ws)
    for row in rows:
        ws.append(list(row))


def _write_eq_sheet(wb, tab_name: str, rows: list):
    ws = wb.create_sheet(tab_name)
    ws.append(["component", "opening", "movement", "closing"])
    _bold_header(ws)
    for row in rows:
        ws.append(list(row))


def _write_ico_sheet(wb, rows: list):
    """Write ICO tab. Columns: account_no | to_party | account_type | amount | currency | description"""
    ws = wb.create_sheet("ICO")
    ws.append(["account_no", "to_party", "account_type", "amount", "currency", "description"])
    _bold_header(ws)
    for row in rows:
        ws.append(list(row))


def _write_meta_sheet(wb, entity_name, currency, reporting_period, ownership_pct):
    ws = wb.create_sheet("META")
    ws.append(["key", "value"])
    _bold_header(ws)
    ws.append(["entity_name", entity_name])
    ws.append(["currency", currency])
    ws.append(["reporting_period", reporting_period])
    ws.append(["ownership_pct", ownership_pct])


def _bold_header(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True)


if __name__ == "__main__":
    main()
