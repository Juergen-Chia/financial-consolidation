from __future__ import annotations
import logging
from pathlib import Path

import openpyxl

from src.models import AccountLine, EntityMeta, EquityLine, SubsidiaryData
from src.coa_mapper import COAMapping

logger = logging.getLogger(__name__)

REQUIRED_TABS = {"BS", "IS", "CF", "EQ", "META"}
CONSECUTIVE_BLANK_STOP = 3
STATEMENT_TABS = {"BS", "IS", "CF"}


class SubsidiaryReadError(Exception):
    pass


def discover_subsidiaries(subsidiaries_dir: Path) -> list[Path]:
    paths = sorted(subsidiaries_dir.glob("*.xlsx"))
    logger.info("Discovered %d subsidiary file(s): %s", len(paths), [p.name for p in paths])
    return paths


def read_subsidiary(path: Path, coa: COAMapping) -> SubsidiaryData:
    logger.info("Reading subsidiary: %s", path.name)
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:
        raise SubsidiaryReadError(f"Cannot open {path.name}: {exc}") from exc

    missing = REQUIRED_TABS - set(wb.sheetnames)
    if missing:
        raise SubsidiaryReadError(f"{path.name} missing tabs: {missing}")

    meta = _read_meta_tab(wb["META"], path.name)
    unmapped: list[dict] = []

    def read_stmt(tab_name: str, statement: str) -> list[AccountLine]:
        rows = _read_statement_rows(wb[tab_name], path.name, tab_name)
        lines: list[AccountLine] = []
        for local_code, account_name, amount in rows:
            group_code = coa.get_group_code(meta.entity_name, local_code)
            if group_code is None:
                unmapped.append({
                    "entity": meta.entity_name,
                    "local_code": local_code,
                    "account_name": account_name,
                    "amount": amount,
                    "statement": statement,
                })
                logger.warning(
                    "Unmapped account '%s' (%s) in %s [%s] — excluded from consolidation",
                    local_code, account_name, meta.entity_name, statement,
                )
                continue
            lines.append(AccountLine(
                group_code=group_code,
                group_account_name=coa.get_group_account_name(group_code),
                amount_local=amount,
                amount_usd=0.0,
                currency=meta.currency,
                entity_name=meta.entity_name,
                statement=statement,
            ))
        return lines

    bs = read_stmt("BS", "BS")
    is_ = read_stmt("IS", "IS")
    cf = read_stmt("CF", "CF")
    eq = _read_eq_tab(wb["EQ"], meta.entity_name, path.name)

    return SubsidiaryData(meta=meta, bs=bs, is_=is_, cf=cf, eq=eq, unmapped_accounts=unmapped)


def _read_meta_tab(ws, filename: str) -> EntityMeta:
    data: dict[str, str] = {}
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0] is not None and row[1] is not None:
            data[str(row[0]).strip().lower()] = str(row[1]).strip()

    required = {"entity_name", "currency", "reporting_period", "ownership_pct"}
    missing = required - set(data.keys())
    if missing:
        raise SubsidiaryReadError(f"{filename} META tab missing keys: {missing}")

    try:
        return EntityMeta(
            entity_name=data["entity_name"],
            currency=data["currency"].upper(),
            reporting_period=data["reporting_period"],
            ownership_pct=float(data["ownership_pct"]),
        )
    except Exception as exc:
        raise SubsidiaryReadError(f"{filename} META tab parse error: {exc}") from exc


def _read_statement_rows(ws, filename: str, tab_name: str) -> list[tuple[str, str, float]]:
    results: list[tuple[str, str, float]] = []
    blank_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        code_val = row[0] if len(row) > 0 else None
        name_val = row[1] if len(row) > 1 else None
        amt_val = row[2] if len(row) > 2 else None

        if code_val is None or str(code_val).strip() == "":
            blank_count += 1
            if blank_count >= CONSECUTIVE_BLANK_STOP:
                break
            continue

        blank_count = 0
        code = str(code_val).strip()
        name = str(name_val).strip() if name_val is not None else ""

        if amt_val is None or str(amt_val).strip() == "":
            amount = 0.0
        else:
            try:
                amount = float(amt_val)
            except (ValueError, TypeError):
                raise SubsidiaryReadError(
                    f"{filename} [{tab_name}] non-numeric amount for account '{code}': {amt_val!r}"
                )

        results.append((code, name, amount))
    return results


def _read_eq_tab(ws, entity_name: str, filename: str) -> list[EquityLine]:
    lines: list[EquityLine] = []
    blank_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        comp_val = row[0] if len(row) > 0 else None
        if comp_val is None or str(comp_val).strip() == "":
            blank_count += 1
            if blank_count >= CONSECUTIVE_BLANK_STOP:
                break
            continue
        blank_count = 0

        def _float(v) -> float:
            if v is None:
                return 0.0
            try:
                return float(v)
            except (ValueError, TypeError):
                return 0.0

        opening = _float(row[1] if len(row) > 1 else None)
        movement = _float(row[2] if len(row) > 2 else None)
        closing = _float(row[3] if len(row) > 3 else None)

        lines.append(EquityLine(
            component=str(comp_val).strip(),
            opening_local=opening,
            movement_local=movement,
            closing_local=closing,
            entity_name=entity_name,
        ))
    return lines
