from __future__ import annotations
import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

from src.models import ConsolidatedData

logger = logging.getLogger(__name__)

_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_BOLD_FONT = Font(bold=True, size=10)
_NORMAL_FONT = Font(size=10)
_ALT_FILL = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
_NUM_FORMAT = '#,##0.00'


class ReportWriter:
    def __init__(self, output_path: Path) -> None:
        self._output_path = output_path

    def write(self, data: ConsolidatedData) -> None:
        output_path = self._output_path
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default sheet

        # Determine entity columns (+ NCI if present) for financial statement tabs
        entity_cols = list(data.entities)
        # Check if any bs_row has NCI key
        has_nci = any("NCI" in row for row in data.bs_rows)
        if has_nci:
            entity_cols = list(data.entities) + ["NCI"]

        _write_statement_tab(wb, "Consolidated BS", data.bs_rows, entity_cols, "group_code", "group_account_name")
        _write_statement_tab(wb, "Consolidated IS", data.is_rows, entity_cols, "group_code", "group_account_name")
        _write_statement_tab(wb, "Consolidated CF", data.cf_rows, entity_cols, "group_code", "group_account_name")
        _write_eq_tab(wb, "Consolidated EQ", data.eq_rows, data.entities)
        _write_eliminations_tab(wb, "Eliminations Log", data.eliminations)
        _write_fx_tab(wb, "FX Translation Log", data.fx_log)
        _write_validation_tab(wb, "Validation", data.validation_messages)

        wb.save(output_path)
        logger.info("Report written to %s", output_path)


def _write_statement_tab(wb, tab_name, rows, entity_cols, code_col, name_col):
    ws = wb.create_sheet(tab_name)
    headers = [code_col.replace("_", " ").title(), name_col.replace("_", " ").title()] + entity_cols + ["Group Total"]
    _write_header_row(ws, headers)

    for i, row in enumerate(rows, start=2):
        ws.cell(i, 1, row.get(code_col, "")).font = _NORMAL_FONT
        ws.cell(i, 2, row.get(name_col, "")).font = _NORMAL_FONT
        for j, entity in enumerate(entity_cols, start=3):
            cell = ws.cell(i, j, row.get(entity, 0.0))
            cell.number_format = _NUM_FORMAT
            cell.font = _NORMAL_FONT
        total_cell = ws.cell(i, 3 + len(entity_cols), row.get("Group Total", 0.0))
        total_cell.number_format = _NUM_FORMAT
        total_cell.font = _BOLD_FONT
        if i % 2 == 0:
            for col in range(1, 3 + len(entity_cols) + 1):
                ws.cell(i, col).fill = _ALT_FILL

    _auto_width(ws, headers)


def _write_eq_tab(wb, tab_name, rows, entities):
    ws = wb.create_sheet(tab_name)
    headers = ["Component", "Period"] + list(entities) + ["Group Total"]
    _write_header_row(ws, headers)

    for i, row in enumerate(rows, start=2):
        ws.cell(i, 1, row.get("component", "")).font = _NORMAL_FONT
        ws.cell(i, 2, row.get("period", "")).font = _NORMAL_FONT
        for j, entity in enumerate(entities, start=3):
            cell = ws.cell(i, j, row.get(entity, 0.0))
            cell.number_format = _NUM_FORMAT
            cell.font = _NORMAL_FONT
        total_cell = ws.cell(i, 3 + len(entities), row.get("Group Total", 0.0))
        total_cell.number_format = _NUM_FORMAT
        total_cell.font = _BOLD_FONT
        if i % 2 == 0:
            for col in range(1, 3 + len(entities) + 1):
                ws.cell(i, col).fill = _ALT_FILL

    _auto_width(ws, headers)


def _write_eliminations_tab(wb, tab_name, eliminations):
    ws = wb.create_sheet(tab_name)
    headers = ["Reference", "Type", "Source", "From Entity", "To Entity", "Account Code", "Account Name", "Statement", "Amount USD", "Description"]
    _write_header_row(ws, headers)

    for i, e in enumerate(eliminations, start=2):
        ws.cell(i, 1, e.reference).font = _NORMAL_FONT
        ws.cell(i, 2, e.elimination_type).font = _NORMAL_FONT
        ws.cell(i, 3, e.source).font = _NORMAL_FONT
        ws.cell(i, 4, e.from_entity).font = _NORMAL_FONT
        ws.cell(i, 5, e.to_entity).font = _NORMAL_FONT
        ws.cell(i, 6, e.account_code).font = _NORMAL_FONT
        ws.cell(i, 7, e.account_name).font = _NORMAL_FONT
        ws.cell(i, 8, e.statement).font = _NORMAL_FONT
        amt_cell = ws.cell(i, 9, e.amount_usd)
        amt_cell.number_format = _NUM_FORMAT
        amt_cell.font = _NORMAL_FONT
        ws.cell(i, 10, e.description).font = _NORMAL_FONT
        if i % 2 == 0:
            for col in range(1, 11):
                ws.cell(i, col).fill = _ALT_FILL

    _auto_width(ws, headers)


def _write_fx_tab(wb, tab_name, fx_log):
    ws = wb.create_sheet(tab_name)
    headers = ["Entity", "Statement", "Group Code", "Account Name", "Amount Local", "Currency", "Rate Type", "Rate Used", "Amount USD"]
    _write_header_row(ws, headers)

    for i, e in enumerate(fx_log, start=2):
        ws.cell(i, 1, e.entity_name).font = _NORMAL_FONT
        ws.cell(i, 2, e.statement).font = _NORMAL_FONT
        ws.cell(i, 3, e.group_code).font = _NORMAL_FONT
        ws.cell(i, 4, e.account_name).font = _NORMAL_FONT
        local_cell = ws.cell(i, 5, e.amount_local)
        local_cell.number_format = _NUM_FORMAT
        local_cell.font = _NORMAL_FONT
        ws.cell(i, 6, e.currency).font = _NORMAL_FONT
        ws.cell(i, 7, e.rate_type).font = _NORMAL_FONT
        rate_cell = ws.cell(i, 8, e.rate_used)
        rate_cell.number_format = "0.0000"
        rate_cell.font = _NORMAL_FONT
        usd_cell = ws.cell(i, 9, e.amount_usd)
        usd_cell.number_format = _NUM_FORMAT
        usd_cell.font = _NORMAL_FONT
        if i % 2 == 0:
            for col in range(1, 10):
                ws.cell(i, col).fill = _ALT_FILL

    _auto_width(ws, headers)


def _write_validation_tab(wb, tab_name, messages):
    ws = wb.create_sheet(tab_name)
    headers = ["Severity", "Category", "Message", "Detail"]
    _write_header_row(ws, headers)

    severity_colors = {"ERROR": "FF0000", "WARNING": "FFC000", "INFO": "70AD47"}

    for i, msg in enumerate(messages, start=2):
        color = severity_colors.get(msg.severity, "000000")
        sev_cell = ws.cell(i, 1, msg.severity)
        sev_cell.font = Font(bold=True, color=color, size=10)
        ws.cell(i, 2, msg.category).font = _NORMAL_FONT
        ws.cell(i, 3, msg.message).font = _NORMAL_FONT
        ws.cell(i, 4, msg.detail).font = _NORMAL_FONT
        if i % 2 == 0:
            for col in range(1, 5):
                ws.cell(i, col).fill = _ALT_FILL

    _auto_width(ws, headers)


def _write_header_row(ws, headers: list[str]) -> None:
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(1, j, h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28


def _auto_width(ws, headers: list[str]) -> None:
    for j, h in enumerate(headers, start=1):
        col_letter = get_column_letter(j)
        max_len = max(len(str(h)), 12)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
